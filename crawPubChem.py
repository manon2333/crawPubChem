from openpyxl import load_workbook

import json
import requests
import re



def getHTMLText1(url,header):
    try:
        r = requests.get(url, header, timeout = 30)
        r.raise_for_status()
        r.encoding = r.apparent_encoding  #防止中文乱码
        #把text内容转为JSON
        json_str=json.loads(r.text)
        # print(json_str)
        # print(type(json_str))
        for keys, values in json_str.items():  # 用序列解包的方法遍历字典中的元素，输出的样式有所变化
          ConceptsAndCIDs=values
          for key, value in ConceptsAndCIDs.items():
              cid = value
              CID = cid[0]
        return CID
    except:
        return ""



def getHTMLText2(url,header,ilt):
    try:
        r = requests.get(url, header, timeout = 30)
        r.raise_for_status()
        r.encoding = r.apparent_encoding  #防止中文乱码
        #把text内容转为JSON
        json_str=json.loads(r.text)
       # print(json_str)
       # print(type(json_str))
        for keys, values in json_str.items():  # 用序列解包的方法遍历字典中的元素，输出的样式有所变化
          for key, value in values.items():  # 用序列解包的方法遍历字典中的元素，输出的样式有所变化
              #print(key)
              #print(value)
              if(key == 'RecordTitle'):
                  englishName=value;
                  # print(englishName)
                  ilt.append(englishName)
              if (key == 'Section'):
                  for item  in  value:
                      #print(item)
                      for ke, va in item.items():
                          #print(ke)
                          #print(va)
                          if(ke == "Section"):
                            #print(va)
                            for ite in va:
                               #print(ite.get('TOCHeading'))
                               if(ite.get('TOCHeading')=="Computed Descriptors"):
                                   section = ite.get("Section")
                                   for i in section:
                                       if(i.get('TOCHeading')=="IUPAC Name"):
                                          information=i.get('Information')
                                          for j in information:
                                              IUPACName =(j.get('Value').get('StringWithMarkup'))[0].get('String')
                                              ilt.append(IUPACName)
                                       if (i.get('TOCHeading') == "Canonical SMILES"):
                                           information = i.get('Information')
                                           for j in information:
                                               SmileName = (j.get('Value').get('StringWithMarkup'))[0].get('String')
                                               ilt.append(SmileName)
        return ilt
    except:
        return ""




def readandwriteExcel():
    wb = load_workbook(r'D:\5.刘赛娃-数据整理-2000化合物出峰情况汇总信息.xlsx')
    ws = wb.active
    sheet_ranges = wb[wb.sheetnames[0]]  # 定位到表格第一张表
    num = 1
    for row in sheet_ranges.rows:  # 循环打印行
        if row[9].value is not None and  row[9].value != "CAS号" and row[9].value!='CAS ':# 判断CAS不为空
            searchNum=row[9].value
            start_url = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/concepts/name/JSON?name=' + searchNum
            header = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36'
            }
            print(start_url)
            infoList = []#存放名称
            #爬取并解析
            CID = getHTMLText1(start_url, header)
            end_url = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/'+str(CID)+'/JSON/'
            getHTMLText2(end_url, header,infoList)
            ws.cell(num, 12, infoList[0])
            ws.cell(num, 13, infoList[2])
            ws.cell(num, 14, infoList[1])
        num = num+1
    wb.save(r'D:\5.刘赛娃-数据整理-2000化合物出峰情况汇总信息.xlsx')



def main():
    readandwriteExcel()


main()



